from Constants import Constants as Const
import os
try:
    import openpyxl
except:
    os.system("python -m pip install openpyxl")
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font,Fill,PatternFill


def getrange(column, row, column2='-1', row2=-1):
    if column2 == '-1' and row2 == -1:
        return column + str(row)
    return column + str(row) + ":" + column2 + str(row2)


class Excel:
    def __init__(self, filename):
        self.__wb = None
        self.__conformationCell = None
        self.__errorCell = None
        self.__sheet = None
        self.__filename = filename
        self.setupexcel()
        print("Excel File Opened")

    def setupexcel(self):
        try:
            try:
                print("trying opening")
                self.__wb = load_workbook(self.__filename)
                print("sucessful open")
            except:
                self.__wb = Workbook()
                self.__save_file()
                self.__wb = load_workbook(self.__filename)
                print("file created")

            self.__sheet = [None, None]
            try:
                self.__sheet[Const.BANK_NIFTY] = self.__wb[Const.BANK_NIFTY_NAME]
                print("got nifty")
            except:
                print("creating nifty")
                self.__sheet[Const.BANK_NIFTY] = self.__wb.create_sheet(title=Const.BANK_NIFTY_NAME, index=2)

            try:
                self.__sheet[Const.NIFTY] = self.__wb[Const.NIFTY_NAME]
                print("got bank nifty")
            except:
                print("creating bank nifty")
                self.__sheet[Const.NIFTY] = self.__wb.create_sheet(title=Const.NIFTY_NAME, index=1)

            self.__save_file()
            try:
                self.__setcolumnnames(index=Const.NIFTY)
                print("posted nifty")
            except:
                print("nifty column fail")
            try:
                self.__setcolumnnames(index=Const.BANK_NIFTY)
                print("posted banknifty")
            except:
                print("banknifty column fail")
            self.__save_file()
            try:
                self.__conformationCell = self.__getconformationcell()
                print("got conformtion")
            except:
                print("failed to get confrm")
            try:
                self.__errorCell = self.__getErrorCell()
                print("got error cells")
            except:
                print("failed to get error cells")
        except Exception as e:
            print("error in test file, ",str(e))
            # self.setupexcel()

    def __getErrorCell(self):
        error = [None, None]
        error[Const.NIFTY] = self.__geterrcells(index=Const.NIFTY)
        error[Const.BANK_NIFTY] = self.__geterrcells(index=Const.BANK_NIFTY)
        return error

    def __geterrcells(self, index):
        row = Const.DIALOGUE_BOX_ROW
        t1 = getrange(column='B', column2='L', row=row, row2=row)
        t2 = self.__sheet[index][getrange(column='B', row=row)]
        t2.font = Font(size=16,color=Const.RED)
        return [t1, t2]

    def __getconformationcell(self):
        cells = [None, None]
        cells[Const.NIFTY] = self.__getconfrmcell(index=Const.NIFTY)
        cells[Const.BANK_NIFTY] = self.__getconfrmcell(index=Const.BANK_NIFTY)
        return cells

    def __getconfrmcell(self, index):
        row = Const.DIALOGUE_BOX_ROW
        column = 'A'
        t = self.__sheet[index][getrange(column=column, row=row)]
        t.font = Font(size=16, color=Const.GREEN,bold=True)
        return t

    def __setconformationcell(self, text):
        self.__conformationCell[Const.NIFTY].value = text
        self.__conformationCell[Const.BANK_NIFTY].value = text

    def __mergecells(self, index, left, right, top, bottom, value):
        self.__sheet[index].merge_cells(getrange(column=left, column2=right, row=top, row2=bottom))
        self.__sheet[index][getrange(row=top, column=left)].value = value

    def __setcolumnnames(self, index):

        try:
            self.__mergecells(index=index, top=1, bottom=1, left='A', right='M', value=Const.WELCOME_TEXT)
        except:
            print("welcome text merging fail")

        # CALLS
        try:
            self.__mergecells(index=index, value='CALLS', top=2, bottom=2, left='A', right='F')
        except:
            print("calls merginf failed")
        # PUTS
        try:
            self.__mergecells(index=index, value='PUTS', top=2, bottom=2, left='H', right='M')
        except:
            print("puts merging failed")

        # LTP
        self.__sheet[index][getrange(column=Const.CALLS_LTP, row=3)].value = Const.LTP
        self.__sheet[index][getrange(column=Const.PUTS_LTP, row=3)].value = Const.LTP

        # OI
        self.__sheet[index][getrange(column=Const.CALLS_OI, row=3)].value = Const.OI
        self.__sheet[index][getrange(column=Const.PUTS_OI, row=3)].value = Const.OI

        # CHANGE OF OI
        self.__sheet[index][getrange(column=Const.CALLS_CHANGE_IN_OI, row=3)].value = Const.CHANGE_IN_OI
        self.__sheet[index][getrange(column=Const.PUTS_CHANGE_IN_OI, row=3)].value = Const.CHANGE_IN_OI

        # TREND 1
        self.__sheet[index][getrange(column=Const.CALLS_TREND1, row=3)].value = Const.getTrends(1)
        self.__sheet[index][getrange(column=Const.PUTS_TREND1, row=3)].value = Const.getTrends(1)

        # TREND 2
        self.__sheet[index][getrange(column=Const.CALLS_TREND2, row=3)].value = Const.getTrends(2)
        self.__sheet[index][getrange(column=Const.PUTS_TREND2, row=3)].value = Const.getTrends(2)

        # TREND 3
        self.__sheet[index][getrange(column=Const.CALLS_TREND3, row=3)].value = Const.getTrends(3)
        self.__sheet[index][getrange(column=Const.PUTS_TREND3, row=3)].value = Const.getTrends(3)

        # STRIKE PRICE
        self.__sheet[index][getrange(column=Const.STRIKE_PRICE_COLUMN, row=3)].value = Const.STRIKE_PRICE

        # SET HEADING STYLE
        for i in 'ABCDEFGHIJKLM':
            for j in range(1,4):
                self.__sheet[index][getrange(column=i, row=j)].font = Font(
                    size=Const.HEADING_SIZE, bold=True)

    def __settimestamp(self, index, price, time, date, marketstatus):
        optionname = Const.NIFTY_NAME if index == Const.NIFTY else Const.BANK_NIFTY_NAME
        self.__sheet[index]['A1'].value = optionname + " : " + str(price) + " as of " + time + " on " + date + (
            ", Market has been closed" if marketstatus == False else "")
        return True

    def __setcellattributes(self, index, colunmname, attribute_data, row):
        t = self.__sheet[index][getrange(column=colunmname, row=row)]
        t.value = attribute_data[Const.TEXT]
        style = attribute_data[Const.FONT_STYLE]
        isbold,isitalic = False,False
        if style == Const.BOLD_ITALIC:
            isbold = True
            isitalic = True
        elif style == Const.BOLD:
            isbold = True
            isitalic = False
        elif style==Const.ITALIC:
            isbold = False
            isitalic = True
        else:
            isbold=False
            isitalic=False

        t.font = Font(size=attribute_data[Const.FONT_SIZE],
                      color=attribute_data[Const.FONT_COLOR],
                      bold=isbold,
                      italic=isitalic)
        if attribute_data[Const.CELL_FILL_COLOR] != None:
            t.fill = PatternFill(fill_type='solid', start_color=attribute_data[Const.CELL_FILL_COLOR],
                                 end_color=attribute_data[Const.CELL_FILL_COLOR])
        else:
            t.fill = PatternFill(fill_type='solid', start_color=Const.WHITE,
                                 end_color=Const.WHITE)

    def __setcalls(self, index, callsdata, row):
        self.__setcellattributes(index=index,
                                 attribute_data=callsdata[Const.OI],
                                 colunmname=Const.CALLS_OI,
                                 row=row)
        self.__setcellattributes(index=index,
                                 attribute_data=callsdata[Const.CHANGE_IN_OI],
                                 colunmname=Const.CALLS_CHANGE_IN_OI,
                                 row=row)
        self.__setcellattributes(index=index,
                                 attribute_data=callsdata[Const.LTP],
                                 colunmname=Const.CALLS_LTP,
                                 row=row)
        self.__setcellattributes(index=index,
                                 attribute_data=callsdata[Const.TRENDS1],
                                 colunmname=Const.CALLS_TREND1,
                                 row=row)
        self.__setcellattributes(index=index,
                                 attribute_data=callsdata[Const.TRENDS2],
                                 colunmname=Const.CALLS_TREND2,
                                 row=row)
        self.__setcellattributes(index=index,
                                 attribute_data=callsdata[Const.TRENDS3],
                                 colunmname=Const.CALLS_TREND3,
                                 row=row)

    def __setputs(self, index, putsdata, row):
        self.__setcellattributes(index=index,
                                 attribute_data=putsdata[Const.OI],
                                 colunmname=Const.PUTS_OI,
                                 row=row)
        self.__setcellattributes(index=index,
                                 attribute_data=putsdata[Const.CHANGE_IN_OI],
                                 colunmname=Const.PUTS_CHANGE_IN_OI,
                                 row=row)
        self.__setcellattributes(index=index,
                                 attribute_data=putsdata[Const.LTP],
                                 colunmname=Const.PUTS_LTP,
                                 row=row)
        self.__setcellattributes(index=index,
                                 attribute_data=putsdata[Const.TRENDS1],
                                 colunmname=Const.PUTS_TREND1,
                                 row=row)
        self.__setcellattributes(index=index,
                                 attribute_data=putsdata[Const.TRENDS2],
                                 colunmname=Const.PUTS_TREND2,
                                 row=row)
        self.__setcellattributes(index=index,
                                 attribute_data=putsdata[Const.TRENDS3],
                                 colunmname=Const.PUTS_TREND3,
                                 row=row)

    def __postdatainexcel(self, index, Data):
        strikes = Data[Const.STRIKES_LIST]
        for i in range(len(strikes)):
            row = i + 4
            data = Data[strikes[i]]
            self.__setcellattributes(index=index,
                                     attribute_data=data[Const.STRIKE_PRICE],
                                     colunmname=Const.STRIKE_PRICE_COLUMN,
                                     row=row)

            self.__setcalls(index=index,
                            row=row,
                            callsdata=data[Const.CALLS])

            self.__setputs(index=index,
                           row=row,
                           putsdata=data[Const.PUTS])

    def __posterror(self, error, index):
        s = ''
        if error == Const.NO_INTERNET:
            s = "Cannot connect to internet, check your connection"
        else:
            s = "Cannot retrieve data from site, site is not working"
        self.__sheet[index].merge_cells(self.__errorCell[index][0])
        self.__errorCell[index][1].value = s

    def __unposterror(self, index):
        try:
            self.__sheet[index].unmerge_cells(self.__errorCell[index][0])
        except:
            print("no error")
        self.__errorCell[index][1].value = ''

    def postinexcel(self, data):
        self.__setconformationcell(text="updating...")
        self.__save_file()
        for index in (Const.NIFTY, Const.BANK_NIFTY):
            optiondata = data[index]
            if optiondata[Const.ERROR] != None:
                self.__setconformationcell(text="")
                self.__posterror(error=optiondata[Const.ERROR], index=index)
                self.__save_file()
                return
            else:
                self.__unposterror(index=index)
            if self.__settimestamp(index=index,
                                   price=optiondata[Const.PRICE],
                                   time=optiondata[Const.TIME],
                                   date=optiondata[Const.DATE],
                                   marketstatus=optiondata[Const.MARKET_STATUS]):
                self.__postdatainexcel(index=index, Data=optiondata)
        self.__setconformationcell(text="")
        self.__save_file()

    def __save_file(self):
        self.__wb.save(filename=self.__filename)
